import openpyxl
import psycopg2
from faker import Faker
fake = Faker()

CONNECTION_PARAMETERS = {
  'dbname': 'postgres',
  'user': 'postgres',
  'password': 'postgrespassword',
  'host': 'localhost',
  'port': '5432'
}

def clear_tables(cur):
  cur.execute("""
    DELETE FROM student_labels;
    DELETE FROM students;
    DELETE FROM admins;
    DELETE FROM colleges;
    DELETE FROM degree_types;
    DELETE FROM enrollment_statuses;
    DELETE FROM genders;
    DELETE FROM highschools;
    DELETE FROM interactions;
    DELETE FROM interaction_types;
    DELETE FROM labels;
    DELETE FROM person;
    DELETE FROM registration_statuses;
  """)

def get_all(ws, *fields):
  ans = []
  for row in ws.iter_rows(min_row=2):
    relation = {}
    for cell in row:
      field = ws.cell(row=1, column=cell.column).value
      if field in fields:
        relation[field] = cell.value
    ans.append(relation)
  return ans

def get_column_number(ws, field_name):
  for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
      if cell.value == field_name:
        return cell.column
  return None

def get_unique(ws, fields, unique):
  result = []
  uniqs = set()
  unique_field_column_number = get_column_number(ws, unique)

  for row in ws.iter_rows(min_row=2):
    unique_field_value = ws.cell(row=row[0].row, column=unique_field_column_number).value
    if (unique_field_value in uniqs or unique_field_value is None):
      continue
    record = {}
    for cell in row:
      field = ws.cell(row=1, column=cell.column).value
      if field in fields:
        record[field] = cell.value
        if field == unique:
          uniqs.add(cell.value)
    result.append(record)
  
  return result

def get_unique_labels(ws):
  uniqs = set()
  label_col_number = get_column_number(ws, 'Labels')
  for row in ws.iter_rows(min_row=2, min_col=label_col_number, max_col=label_col_number):
    cell = row[0]
    labels = cell.value.split("; ")
    for label in labels:
      uniqs.add((label.strip()))
  return list(uniqs)

def main():
  # Open Gradsnapp data
  wb = openpyxl.load_workbook(
    filename='fwddata/Gradsnapp Data - Cleaned of Personal Identifable Info (No Phone Email).xlsx',
    data_only=True
  )
  ws = wb.active

  # Connect to DB
  with psycopg2.connect(**CONNECTION_PARAMETERS) as conn:
    with conn.cursor() as cur:
      clear_tables(cur)

      students = get_all(
        ws,
        'Student Id',
        'Year',
        'Status',
        'FirstName',
        'LastName',
        'Gender',
        'Labels',
        'Email',
        'Phone',
        'AcademicScore',
        'RegistrationStatus',
        'PostHsPlans',
        'Major',
        'PlannedDegreeType',
        'ExpectedGraduationYear',
        'CollegeName',
        'HighSchoolName'
      )
      
      # insert highschools
      highschools = get_unique(
        ws, fields=['HighSchoolName','HighSchoolState'], unique='HighSchoolName'
      )
      highschool_id = 1
      for highschool in highschools:
        highschool['highschool_id'] = highschool_id
        cur.execute("""
          INSERT INTO highschools (id, name, state)
          VALUES (%s, %s, %s)
        """, (
          highschool['highschool_id'],
          highschool['HighSchoolName'],
          highschool['HighSchoolState']
        ))
        highschool_id += 1
        for student in students:
          if student['HighSchoolName'] == highschool['HighSchoolName']:
            student['highschool_id'] = highschool['highschool_id']

      # insert colleges
      colleges = get_unique(
        ws,
        fields=['CollegeName','CollegeCity','CollegeState'],
        unique='CollegeName'
      )
      college_id = 1
      for college in colleges:
        college['college_id'] = college_id
        cur.execute("""
          INSERT INTO colleges (id, name, city, state)
          VALUES (%s, %s, %s, %s)
        """, (
          college['college_id'],
          college['CollegeName'],
          college['CollegeCity'],
          college['CollegeState']
        ))
        college_id += 1 
        for student in students:
          if student['CollegeName'] == college['CollegeName']:
            student['college_id'] = college['college_id']

      # insert genders
      genders = get_unique(ws, fields=['Gender'], unique='Gender')
      for gender in genders:
        cur.execute("""
          INSERT INTO genders (type)
          VALUES (%s)
        """, (gender['Gender'],))

      # insert labels
      labels = get_unique_labels(ws)
      for label in labels:
        cur.execute("""
          INSERT INTO labels (type)
          VALUES (%s)
        """, (label,))

      # insert degree types
      degree_types = get_unique(
        ws, fields=['PlannedDegreeType'], unique='PlannedDegreeType'
      )
      for degree_type in degree_types:
        cur.execute("""
          INSERT INTO degree_types
          VALUES (%s)
        """,
        (degree_type['PlannedDegreeType'],))

      # insert enrollment statuses
      enrollment_statuses = get_unique(ws, fields=['Status'], unique='Status')
      for status in enrollment_statuses:
        cur.execute("""
          INSERT INTO enrollment_statuses
          VALUES (%s)
        """,
        (status['Status'],))

      # insert registration statuses
      registration_statuses = get_unique(
        ws,
        fields=['RegistrationStatus'],
        unique='RegistrationStatus'
      )
      for reg_status in registration_statuses:
        cur.execute("""
          INSERT INTO registration_statuses
          VALUES (%s)
        """,
        (reg_status['RegistrationStatus'],))

      # insert students into person table
      person_id = 1
      for student in students:
        student['person_id'] = person_id
        cur.execute("""
          INSERT INTO person (id, first_name, last_name, email)
          VALUES (%s, %s, %s, %s)
        """, (
          student['person_id'],
          student['FirstName'] or fake.first_name() + ' (FAKE)',
          student['LastName'] or fake.last_name() + ' (FAKE)',
          student['Email'] or fake.email() + ' (FAKE)'
        ))
        person_id += 1
        

      # insert students into students table
      for student in students:
        if student['HighSchoolName'] is None:
          student['highschool_id'] = None
        if student['CollegeName'] is None:
          student['college_id'] = None
        # handle one student ID that contains a letter
        if student['Student Id'][0].isalpha():
          student['Student Id'] = 10000000
        # convert majors that are empty strings to None
        if not student['Major']:
          student['Major'] = None
        cur.execute("""
          INSERT INTO students (id, person_id, year, enrollment_status, gender,
                                phone, highschool_id, college_id,
                                hs_academic_score, post_hs_plans,
                                registration_status, major, degree_type,
                                expected_graduation_year)
          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
          student['Student Id'],
          student['person_id'],
          student['Year'],
          student['Status'],
          student['Gender'],
          student['Phone'],
          student['highschool_id'],
          student['college_id'],
          student['AcademicScore'],
          student['PostHsPlans'],
          student['RegistrationStatus'],
          student['Major'],
          student['PlannedDegreeType'],
          student['ExpectedGraduationYear']
        ))

      # insert student_labels
      student_label_id = 1
      for student in students:
        labels = [label.strip() for label in student['Labels'].split('; ')]
        for label in labels:
          cur.execute("""
            INSERT INTO student_labels (id, student_id, label_type)
            VALUES (%s, %s, %s)
          """, (
            student_label_id,
            student['Student Id'],
            label
          ))
          student_label_id += 1

      # open interactions worksheet
      wb = openpyxl.load_workbook(
        filename='fwddata/PersistEngagementReport_Example.xlsx',
        data_only=True
      )
      ws = wb.active

      # insert admins
      admins = get_unique(
        ws,
        fields=['Created By'],
        unique='Created By'
      )
      # remove NULL admin
      admins = [admin for admin in admins if admin['Created By']] 
      admin_id = 1
      for admin in admins:
        names = admin['Created By'].split(" ")
        admin['first_name'] = names[0]
        admin['last_name'] = names[1]
        admin['password'] = 'password'
        admin['person_id'] = person_id
        admin['id'] = admin_id
        cur.execute("""
          INSERT INTO person (id, first_name, last_name, email)
          VALUES (%s, %s, %s, %s)
        """, (
          admin['person_id'],
          admin['first_name'],
          admin['last_name'],
          admin['email'] if 'email' in admin else fake.email() + ' (FAKE)',
        ))
        cur.execute("""
          INSERT INTO admins (id, person_id, password)
          VALUES (%s, %s, %s)
        """, (
          admin['id'],
          admin['person_id'],
          admin['password']
        ))
        person_id += 1
        admin_id += 1

      # insert interaction_types
      interaction_types = get_unique(
        ws,
        fields=['Interaction Type'],
        unique='Interaction Type'
      )
      for inter_type in interaction_types:
        cur.execute("""
          INSERT INTO interaction_types (type)
          VALUES (%s)
        """, (inter_type['Interaction Type'],))

      """
        Interaction content:
        -- Bulk Email: None
        -- Bulk Sms: SMS Message
        -- Sms Received: SMS Message
        -- Note: Contact Note
      """

      interactions = get_all(ws, 'Interaction Type', 'Student ID', 'Created By',
                            'Created Date', 'SMS Message', 'Contact Note')
                  
      # set interaction content based on type
      for interaction in interactions:
        if interaction['Interaction Type'] == 'Bulk Email':
          interaction['content'] = None
        elif interaction['Interaction Type'] == 'Note':
          interaction['content'] = interaction['Contact Note']
        else:
          interaction['content'] = interaction['SMS Message']

      for interaction in interactions:
          # get person_ids of students in interactions
          interaction['student_person_id'] = None
          for student in students:
            if student['Student Id'] == interaction['Student ID']:
              interaction['student_person_id'] = student['person_id']

          get_admin_person_id_query = """
            SELECT id FROM person
            WHERE concat(person.first_name, ' ', person.last_name) = (%s)
          """

          # set interaction creator and receiver based on type
          if interaction['Interaction Type'] == 'Sms Received':
            interaction['recipient_id'] = None # will have to be replaced with admin's person_id
            interaction['created_by_id'] = interaction['student_person_id']
          else:
            interaction['recipient_id'] = interaction['student_person_id']

            cur.execute(get_admin_person_id_query, (interaction['Created By'],))
            interaction['created_by_id'] = cur.fetchone()[0] # will have to be replaced with admin's person_id

      interaction_id = 1
      for interaction in interactions:
        cur.execute("""
          INSERT INTO interactions (id, interaction_type, created_by_id, recipient_id, date, content)
          VALUES (%s, %s, %s, %s, %s, %s)
        """, (
          interaction_id,
          interaction['Interaction Type'],
          interaction['created_by_id'],
          interaction['recipient_id'],
          interaction['Created Date'],
          interaction['content']
        ))
        interaction_id += 1

if __name__ == '__main__':
  main()
