from psycopg2 import sql, connect
import openpyxl
from faker import Faker
fake = Faker()

CONNECTION_PARAMETERS = {
  'dbname': 'postgres',
  'user': 'postgres',
  'password': 'postgrespassword',
  'host': 'localhost',
  'port': '5432'
}

def clear_tables(cur):
  cur.execute("""DELETE FROM student_labels; DELETE FROM students;
    DELETE FROM admins; DELETE FROM colleges; DELETE FROM degree_types;
    DELETE FROM enrollment_statuses; DELETE FROM genders;
    DELETE FROM highschools; DELETE FROM interactions;
    DELETE FROM interaction_types; DELETE FROM labels; DELETE FROM person;
    DELETE FROM registration_statuses;
  """)

def get_types(rows, column_header):
  types = set()
  for row in rows:
    value = row[column_header]
    if bool(value):
      types.add(value)
  return [(t,) for t in types]

def get(rows, column_headers, unique_header):
  uniqs = set()
  records = []
  for row in rows:
    if row[unique_header] not in uniqs and row[unique_header] is not None:
      uniqs.add(row[unique_header])
      records.append({ header: row[header] for header in column_headers})
  return records

def transfer_students(students, cur):
  person_id = 1
  label_id = 1
  for student in students:
    # insert into person table
    student['person_id'] = person_id
    person_id += 1
    insert(cur, 'person', ['id', 'first_name', 'last_name', 'email'], [(
      student['person_id'],
      '(FAKE) ' + fake.first_name(),
      '(FAKE) ' + fake.last_name(),
      '(FAKE) ' + fake.email()
    )])
    
    # get highschool_id
    cur.execute("""
      SELECT id from highschools
      WHERE name = %s
    """, (student['HighSchoolName'],))
    student['highschool_id'] = cur.fetchone()[0]

    # get college_id
    cur.execute("""
      SELECT id from colleges
      WHERE name = %s
    """, (student['CollegeName'],))
    res = cur.fetchone()
    student['college_id'] = None if res is None else res[0]

    # change if of student with alpha char
    if not student['Student Id'].isnumeric():
      student['Student Id'] = 100000000

    # replace whitespace majors with None  
    if student['Major'] is not None and not student['Major'].strip():
      student['Major'] = None

    # insert into students table
    insert(cur, 'students',
    fields=['id', 'person_id', 'year', 'enrollment_status',
            'gender', 'phone', 'highschool_id', 'college_id', 'hs_academic_score',
            'post_hs_plans', 'registration_status', 'major', 'degree_type',
            'expected_graduation_year'],
    values=[[student['Student Id'], student['person_id'], student['Year'],
            student['Status'], student['Gender'], student['Phone'],
            student['highschool_id'], student['college_id'],
            student['AcademicScore'], student['PostHsPlans'],
            student['RegistrationStatus'], student['Major'],
            student['PlannedDegreeType'], student['ExpectedGraduationYear']]]
    )

    # insert into student_labels table
    labels = student['Labels'].split('; ')
    labels = [label.strip() for label in labels]
    for label in labels:
      insert(cur, 'student_labels', ['id', 'student_id', 'label_type'], [(
        label_id,
        student['Student Id'],
        label
      )])
      label_id += 1

def transfer_admins(interactions, cur):
  admin_names = get_types(interactions, 'Created By')
  admins = []
  for name_tup in admin_names:
    first_name, last_name = name_tup[0].split(" ")
    admins.append({
      'first_name': first_name,
      'last_name': last_name
    })
  add_ids_to(admins)
  person_id = get_next_id(cur, 'person')
  for admin in admins:
    admin['person_id'] = person_id
    person_id += 1
    insert(cur, 'person', ['id', 'first_name', 'last_name', 'email'], [(
      admin['person_id'],
      admin['first_name'],
      admin['last_name'],
      '(FAKE) ' + fake.email()
    )])
    insert(cur, 'admins', ['id', 'person_id', 'password'], [(
      admin['id'],
      admin['person_id'],
      'password'
    )])
  

def get_next_id(cur, table):
  query = sql.SQL("""
    SELECT id
    FROM {}
    ORDER BY id DESC
    LIMIT 1
  """).format(sql.Identifier(table))
  cur.execute(query)
  res = cur.fetchone()
  return res[0] + 1 if res else 1

def transfer_type(students, cur, column_header, table):
  types = get_types(students, column_header)
  insert(cur, table, ['type'], types)

def transfer_labels(students, cur):
  labels_strings = get_types(students, 'Labels')
  result = set()
  for tup in labels_strings:
    labs = tup[0].split("; ")
    for lab in labs:
      result.add(lab.strip())
  values = [(label,) for label in result]
  insert(cur, 'labels', ['type'], values)

def insert(cur, table, fields, values):
  query = sql.SQL("""
    INSERT INTO {table} ({fields})
    VALUES ({placeholders})
  """).format(
    table=sql.Identifier(table),
    fields=sql.SQL(", ").join(map(sql.Identifier, fields)),
    placeholders=sql.SQL(', ').join(sql.Placeholder() * len(fields))
  )
  cur.executemany(query, values)

def add_ids_to(records):
  id = 1
  for record in records:
    record['id'] = id
    id += 1
  return records

def transfer_no_foreign_keys(rows, cur, column_names, unique_column_name,
  table, fields):
  records = get(rows, column_names, unique_column_name)
  add_ids_to(records)
  values = [
    (record['id'],) + tuple(record[col_name] for col_name in column_names)
    for record in records
  ]
  insert(cur, table, fields, values)

def get_data_from_wb(wb_path):
  wb = openpyxl.load_workbook(
    filename=wb_path,
    data_only=True
  )
  ws = wb.active

  # build list of objects
  headers = [cell.value for cell in ws['1']]
  return [
    {attribute: cell.value for attribute, cell in zip(headers, row)}
    for row in ws.iter_rows(min_row=2)
  ]

def get_admin_id(cur, name):
  query = """
    SELECT id FROM person
    WHERE concat(person.first_name, ' ', person.last_name) = (%s)
  """
  cur.execute(query, (name,))
  res = cur.fetchone()
  if res:
    return res[0]
  raise Exception(f"Couldn't find {name} in person table")

def get_person_id(cur, student_id):
  query = """
    SELECT person_id FROM students
    WHERE id = %s
  """
  cur.execute(query, (student_id,))
  res = cur.fetchone()
  return res[0] if res else None

def transfer_interactions(interactions, cur):
  for interaction in interactions:
    admin_name = interaction['Created By']
    if not interaction['Student ID'].isnumeric():
      interaction['Student ID'] = 100000000
    if admin_name:
      interaction['created_by_id'] = get_admin_id(cur, admin_name)
      interaction['recipient_id'] = get_person_id(cur, interaction['Student ID'])
    else:
      interaction['created_by_id'] = get_person_id(cur, interaction['Student ID'])
      interaction['recipient_id'] = None
    
    if interaction['Interaction Type'] == 'Bulk Email':
      interaction['content'] = None
    elif interaction['Interaction Type'] == 'Note':
      interaction['content'] = interaction['Contact Note']
    else:
      interaction['content'] = interaction['SMS Message']

  add_ids_to(interactions)

  for interaction in interactions:
    insert(cur, 'interactions', 
    ['id', 'interaction_type', 'created_by_id', 'recipient_id', 'date', 'content'],
    [(
      interaction['id'],
      interaction['Interaction Type'],
      interaction['created_by_id'],
      interaction['recipient_id'],
      interaction['Created Date'],
      interaction['content']
    )])

if __name__ == '__main__':
  # get data from Excel workbooks
  students = get_data_from_wb(
    'fwddata/Gradsnapp Data - Cleaned of Personal Identifable Info (No Phone Email).xlsx')
  interactions = get_data_from_wb(
    'fwddata/PersistEngagementReport_Example.xlsx')

  # connect to postgres db
  with connect(**CONNECTION_PARAMETERS) as conn:
    with conn.cursor() as cur:
      clear_tables(cur)
      
      # Insert all tables who have only a 'type' column
      transfer_type(students, cur, 'Gender', 'genders')
      transfer_type(students, cur, 'Status', 'enrollment_statuses')
      transfer_type(students, cur, 'PlannedDegreeType', 'degree_types')
      transfer_type(students, cur, 'RegistrationStatus', 'registration_statuses')
      transfer_labels(students, cur)

      # Insert tables that don't have foreign keys
      transfer_no_foreign_keys(students, cur, 
        column_names=['HighSchoolName', 'HighSchoolState'],
        unique_column_name='HighSchoolName',
        table='highschools',
        fields=['id', 'name', 'state'])
      transfer_no_foreign_keys(students, cur, 
        column_names=['CollegeName', 'CollegeCity', 'CollegeState'],
        unique_column_name='CollegeName',
        table='colleges',
        fields=['id', 'name', 'city', 'state'])

      # Insert students 
      transfer_students(students, cur)

      # Insert admins
      transfer_admins(interactions, cur)

      # Insert interaction types
      transfer_type(interactions, cur, 'Interaction Type', 'interaction_types')

      # Insert interactions
      transfer_interactions(interactions, cur)
      